from openpyxl import load_workbook
import json

finalObj = {
    "events": []
}

# Loading the headers
# TODO: There's gotta be some way to optimize this, right?
wb = load_workbook(filename = 'mkesocialism_timeline.xlsx')
sheet = wb['Sheet1']

startDateHeaders = ["year", "month", "day", "hour", "minute", "second", "millisecond", "format"]
# for row in sheet.iter_rows(min_row=1, max_col=4, max_row=1):
#     for cell in row:
#         startDateHeaders.append(cell.value)

endDateHeaders = ["year", "month", "day", "hour", "minute", "second", "millisecond", "format"]
# for row in sheet.iter_rows(min_row=1, min_col=5, max_col=8, max_row=1):
#     for cell in row:
#         endDateHeaders.append(cell.value)

mediaHeaders = ["caption", "credit", "thumb", "url"]
# for row in sheet.iter_rows(min_row=1, min_col=12, max_col=15, max_row=1):
#     for cell in row:
#         mediaHeaders.append(cell.value)

textHeaders = ["headline", "text"]
# for row in sheet.iter_rows(min_row=1, min_col=10, max_col=11, max_row=1):
#     for cell in row:
#         textHeaders.append(cell.value)

numEntries = 0
for row in sheet.iter_cols(min_row=2, min_col=1, max_col=1):
    for cell in row:
        if cell.value is not None:
            numEntries += 1
print(numEntries)

for x in range(0, numEntries):
    tempObj = {}

    startDateData = []
    for row in sheet.iter_rows(min_row=x+2, max_row=x+2, max_col=4):
        for cell in row:
            startDateData.append(cell.value)
        for x in range(0, 3):
            startDateData.append(None)
    endDateData = []
    for row in sheet.iter_rows(min_row=x+2, max_row=x+2, min_col=5, max_col=8):
        for cell in row:
            endDateData.append(cell.value)
        for x in range(0, 3):
            endDateData.append(None)
    mediaData = []
    for row in sheet.iter_rows(min_row=x+2, max_row=x+2, min_col=12, max_col=15):
        for cell in row:
            mediaData.append(cell.value)
    textData = []
    for row in sheet.iter_rows(min_row=x+2, max_row=x+2, min_col=10, max_col=11):
        for cell in row:
            textData.append(cell.value)

    startDateTempObj = {}
    for x in range(0, startDateData.__len__()):
        if startDateData[x] is not None:
            startDateTempObj[startDateHeaders[x]] = startDateData[x]
        else:
            startDateTempObj[startDateHeaders[x]] = ""
    endDateTempObj = {}
    for x in range(0, endDateData.__len__()):
        if endDateData[x] is not None:
            endDateTempObj[endDateHeaders[x]] = endDateData[x]
        else:
            endDateTempObj[endDateHeaders[x]] = ""
    mediaTempObj = {}
    for x in range(0, mediaData.__len__()):
        if mediaData[x] is not None:
            if(mediaHeaders[x] == "url"):
                mediaTempObj[mediaHeaders[x]] = "img/" + mediaData[x]
            else:
                mediaTempObj[mediaHeaders[x]] = mediaData[x]
        else:
            mediaTempObj[mediaHeaders[x]] = ""
    textTempObj = {}
    for x in range(0, textData.__len__()):
        if textData[x] is not None:
            textTempObj[textHeaders[x]] = textData[x]
        else:
            textTempObj[textHeaders[x]] = ""

    tempObj = {
        "start_date": startDateTempObj,
        "end_date": endDateTempObj,
        "media": mediaTempObj,
        "text": textTempObj
    }
    finalObj["events"].append(tempObj)

jsonObj = json.dumps(finalObj, indent=4)
with open('socialism_result.json', 'w') as f:
    f.write(jsonObj)

