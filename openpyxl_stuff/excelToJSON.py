from openpyxl import load_workbook
import json

finalObj = {
    "events": []
}

# Loading the headers
# TODO: There's gotta be some way to optimize this, right?
wb = load_workbook(filename = 'openpyxl_stuff/new2.xlsx')
sheet = wb['Sheet1']

startDateHeaders = ["year", "month", "day", "hour", "minute", "second", "millisecond", "format"]
# for row in sheet.iter_rows(min_row=1, max_col=4, max_row=1):
#     for cell in row:
#         startDateHeaders.append(cell.value)

endDateHeaders = ["year", "month", "day", "hour", "minute", "second", "millisecond", "format"]
# for row in sheet.iter_rows(min_row=1, min_col=5, max_col=8, max_row=1):
#     for cell in row:
#         endDateHeaders.append(cell.value)

mediaHeaders = ["caption", "credit", "thumb", "url"]
# for row in sheet.iter_rows(min_row=1, min_col=12, max_col=15, max_row=1):
#     for cell in row:
#         mediaHeaders.append(cell.value)

textHeaders = ["headline", "text"]
# for row in sheet.iter_rows(min_row=1, min_col=10, max_col=11, max_row=1):
#     for cell in row:
#         textHeaders.append(cell.value)

numEntries = 0
for row in sheet.iter_cols(min_row=2, min_col=1, max_col=1):
    for cell in row:
        if cell.value is not None:
            numEntries += 1
print(str(numEntries) + " entries detected.")

for x in range(0, numEntries):
    tempObj = {}

    startDateData = []
    for row in sheet.iter_rows(min_row=x+2, max_row=x+2, max_col=4):
        for cell in row:
            startDateData.append(cell.value)
        for y in range(0, 3):
            startDateData.append(None)
        y=0
    endDateData = []
    for row in sheet.iter_rows(min_row=x+2, max_row=x+2, min_col=5, max_col=8):
        for cell in row:
            endDateData.append(cell.value)
        for y in range(0, 3):
            endDateData.append(None)
        y=0
    mediaData = []
    for row in sheet.iter_rows(min_row=x+2, max_row=x+2, min_col=12, max_col=15):
        for cell in row:
            mediaData.append(cell.value)
    textData = []
    for row in sheet.iter_rows(min_row=x+2, max_row=x+2, min_col=10, max_col=11):
        for cell in row:
            textData.append(cell.value)

    startDateTempObj = {}
    for z in range(0, startDateData.__len__()):
        if startDateData[z] is not None:
            startDateTempObj[startDateHeaders[z]] = str(startDateData[z])
        else:
            startDateTempObj[startDateHeaders[z]] = ""
    z=0
    endDateTempObj = {}
    for z in range(0, endDateData.__len__()):
        if endDateData[z] is not None:
            endDateTempObj[endDateHeaders[z]] = str(endDateData[z])
        else:
            endDateTempObj[endDateHeaders[z]] = ""
    z=0
    mediaTempObj = {}
    for z in range(0, mediaData.__len__()):
        if mediaData[z] is not None and mediaData[z] != '?' and mediaData[z] != 'n/a' and mediaData[z] != '' and mediaData[z] != ' ' and mediaData[2] is not None and mediaData[2] != '?' and mediaData[2] != 'n/a' and mediaData[2] != '' and mediaData[2] != ' ':
            if mediaHeaders[z] == "url":
                mediaTempObj[mediaHeaders[z]] = "img/" + mediaData[z]
            elif mediaHeaders[z] == "caption":
                if mediaData[z+2] != '?' and mediaData[z+2] != 'n/a' and mediaData[z+2] != '' and mediaData[z+2] != ' ':
                    mediaTempObj[mediaHeaders[z]] = "<a href='" + mediaData[z] + "' target='_blank'>" + mediaData[z+2] + "</a>"
            else:
                mediaTempObj[mediaHeaders[z]] = mediaData[z]
        else:
            mediaTempObj[mediaHeaders[z]] = ""
    
    z=0
    textTempObj = {}
    for z in range(0, textData.__len__()):
        if textData[z] is not None:
            textTempObj[textHeaders[z]] = textData[z]
        else:
            textTempObj[textHeaders[z]] = ""
    z=0

    # A check to make sure that there actually is data for the end date
    blank = True
    for z in range(0, endDateData.__len__()):
        if endDateData[z] != None:
            blank = False
            break
    if blank:
        tempObj = {
            "start_date": startDateTempObj,
            "media": mediaTempObj,
            "text": textTempObj
        }
    else:
        tempObj = {
            "start_date": startDateTempObj,
            "end_date": endDateTempObj,
            "media": mediaTempObj,
            "text": textTempObj
        }

    finalObj["events"].append(tempObj)

jsonObj = json.dumps(finalObj, indent=4)
with open('result_2.json', 'w') as f:
    f.write(jsonObj)
print("JSON file saved.")

