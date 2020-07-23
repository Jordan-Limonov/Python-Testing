from openpyxl import load_workbook
import json

# Loading the headers
# TODO: There's gotta be some way to optimize this, right?
wb = load_workbook(filename = 'test.xlsx')
sheet = wb['od1']

startDateHeaders = []
for row in sheet.iter_rows(min_row=1, max_col=4, max_row=1):
    for cell in row:
        startDateHeaders.append(cell.value)

endDateHeaders = []
for row in sheet.iter_rows(min_row=1, min_col=5, max_col=8, max_row=1):
    for cell in row:
        endDateHeaders.append(cell.value)

mediaHeaders = []
for row in sheet.iter_rows(min_row=1, min_col=12, max_col=15, max_row=1):
    for cell in row:
        mediaHeaders.append(cell.value)

textHeaders = []
for row in sheet.iter_rows(min_row=1, min_col=10, max_col=11, max_row=1):
    for cell in row:
        textHeaders.append(cell.value)

# Creating dictionaries with the headers
# TODO: There's gotta be some way to optimize this, right?
startDateObj = {}
for header in startDateHeaders:
    startDateObj[header] = ""
# print(startDateObj)

endDateObj = {}
for header in endDateHeaders:
    endDateObj[header] = ""
# print(endDateObj)

mediaObj = {}
for header in mediaHeaders:
    mediaObj[header] = ""
# print(mediaObj)

textObj = {}
for header in textHeaders:
    textObj[header] = ""
# print(textObj)

# Creating a final dictionary and converting it to JSON
finalObj = {
    "start_date": startDateObj,
    "end_date": endDateObj,
    "media": mediaObj,
    "text": textObj
}
# Could use this JSON object as a template, copying it for each entry into a temporary variable.
# Everything is supposed to be in one file anyways, so that should work if I just add each entry into a dictionary once I'm done with it.
jsonObj = json.dumps(finalObj, indent=4)
# print(jsonObj)

# Saving that JSON Object as a file
with open('result.json', 'w') as f:
    f.write(jsonObj)

# Let's iterate through everything
# Reminder: You can just use 'is not None' for the equivalent of '!= null'
for row in sheet.values:
    for value in row:
        if value is not None:
            print(value)